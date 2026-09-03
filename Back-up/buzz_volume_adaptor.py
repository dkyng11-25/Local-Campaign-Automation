import os
import copy
import json
import re
import pandas as pd
import requests

from dotenv import load_dotenv
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from enum import Enum
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

# =============================================================================
# 프로젝트 경로 설정
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_ROOT / "output"
PAYLOAD_DIR = PROJECT_ROOT / "payload"
RESPONSE_SAMPLE_DIR = PROJECT_ROOT / "sprinklr_response_samples"

HEADER_ROW = 4
ROW_ID_COLUMN_NAME = "#"
BUZZ_VOLUME_COLUMN_NAME = "Buzz Volume"
SAVE_FAILED_RESPONSES = True

"""사용자 설정 필요"""
SPRINKLR_BASE_URL = "https://api3.sprinklr.com/prod"
ENDPOINT = "/api/v2/reports/query"

# 보안상 실제 값은 코드에 직접 쓰기보다 환경변수로 관리
load_dotenv(
    dotenv_path=PROJECT_ROOT/".env"
)
API_KEY = os.getenv("SPRINKLR_API_KEY")
ACCESS_TOKEN = os.getenv("SPRINKLR_ACCESS_TOKEN")

# =============================================================================
# Excel 경로 생성
# =============================================================================

def build_excel_paths(
    input_date: str,
) -> tuple[Path, Path]:
    """
    YYYY-MM-DD 형식의 사용자 입력을 YYMMDD로 변환한 뒤,
    output/{YYMMDD}/ 폴더에서 LLM 분석 완료 Excel 파일을 찾는다.
    """
    input_date = input_date.strip()

    try:
        parsed_date = datetime.strptime(
            input_date,
            "%Y-%m-%d",
        )
    except ValueError as exc:
        raise ValueError(
            "날짜는 YYYY-MM-DD 형식으로 입력해야 합니다. "
            "예시) 2026-07-14"
        ) from exc

    date_folder_name = parsed_date.strftime("%y%m%d")
    month = parsed_date.month

    date_output_dir = OUTPUT_DIR / date_folder_name

    if not date_output_dir.exists():
        raise FileNotFoundError(
            "해당 날짜의 output 폴더를 찾을 수 없습니다.\n"
            f"사용자 입력: {input_date}\n"
            f"변환된 폴더명: {date_folder_name}\n"
            f"확인 경로: {date_output_dir}"
        )

    if not date_output_dir.is_dir():
        raise NotADirectoryError(
            "해당 경로가 폴더가 아닙니다.\n"
            f"확인 경로: {date_output_dir}"
        )

    input_excel_path = (
        date_output_dir
        / (
            f"{date_folder_name}"
            "_SLCC_SOV_Local Campaign Tracking_"
            f"{month}월_v01.xlsx"
        )
    )

    # 입력 파일명에 suffix만 추가하여 결과 파일명을 생성한다.
    output_excel_path = input_excel_path.with_name(
        f"{input_excel_path.stem}_mentions_updated.xlsx"
    )

    if not input_excel_path.exists():
        raise FileNotFoundError(
            "Mention 업데이트에 사용할 Excel 파일을 "
            "찾을 수 없습니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if not input_excel_path.is_file():
        raise FileNotFoundError(
            "입력 Excel 경로가 파일이 아닙니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    return input_excel_path, output_excel_path


# =============================================================================
# Base payload 로드
# =============================================================================

def load_base_payload(
    payload_path: Path,
) -> dict[str, Any]:
    if not payload_path.exists():
        raise FileNotFoundError(
            f"Base payload 파일을 찾을 수 없습니다: {payload_path}"
        )

    if not payload_path.is_file():
        raise FileNotFoundError(
            f"Base payload 경로가 파일이 아닙니다: {payload_path}"
        )

    try:
        with payload_path.open(
            mode="r",
            encoding="utf-8",
        ) as file:
            payload = json.load(file)

    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Base payload가 올바른 JSON 형식이 아닙니다: {payload_path}"
        ) from exc

    if not isinstance(payload, dict):
        raise TypeError(
            "Base payload의 최상위 구조는 JSON object여야 합니다."
        )

    return payload


# =============================================================================
# 날짜 data cut 설정
# =============================================================================

SEOUL_TIMEZONE = ZoneInfo("Asia/Seoul")

# Daily 데이터 컷의 고정 시작일
DAILY_START_DATE = date(2026, 7, 1)

# 마지막 초 전체를 포함하도록 999 milliseconds까지 설정
DAILY_END_TIME = time(
    hour=18,
    minute=0,
    second=59,
    microsecond=999_000,
)

WEEKLY_START_TIME = time(
    hour=0,
    minute=0,
    second=0,
    microsecond=0,
)

WEEKLY_END_TIME = time(
    hour=23,
    minute=59,
    second=59,
    microsecond=999_000,
)


class DataCutType(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"


@dataclass(frozen=True)
class DateTimeRange:
    data_cut_type: DataCutType
    reference_date: date
    start_datetime: datetime
    end_datetime: datetime


def parse_reference_date(
    value: str,
) -> date:
    value = value.strip()

    try:
        return datetime.strptime(
            value,
            "%Y-%m-%d",
        ).date()

    except ValueError as exc:
        raise ValueError(
            "기준 날짜는 YYYY-MM-DD 형식이어야 합니다. "
            f"입력값: {value!r}"
        ) from exc


def build_daily_range(
    reference_date: date,
) -> DateTimeRange:
    if reference_date < DAILY_START_DATE:
        raise ValueError(
            "Daily 기준 날짜는 고정 시작일보다 "
            "이전일 수 없습니다. "
            f"고정 시작일: {DAILY_START_DATE}, "
            f"입력 날짜: {reference_date}"
        )

    start_datetime = datetime.combine(
        DAILY_START_DATE,
        time.min,
        tzinfo=SEOUL_TIMEZONE,
    )

    end_datetime = datetime.combine(
        reference_date,
        DAILY_END_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    return DateTimeRange(
        data_cut_type=DataCutType.DAILY,
        reference_date=reference_date,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
    )


def build_weekly_range(
    reference_date: date,
) -> DateTimeRange:
    # Python weekday(): 월요일=0, 일요일=6
    if reference_date.weekday() != 0:
        raise ValueError(
            "Weekly 기준 날짜는 반드시 월요일이어야 합니다. "
            f"입력 날짜: {reference_date}, "
            f"요일 번호: {reference_date.weekday()}"
        )

    previous_monday = reference_date - timedelta(days=7)
    previous_sunday = reference_date - timedelta(days=1)

    start_datetime = datetime.combine(
        previous_monday,
        WEEKLY_START_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    end_datetime = datetime.combine(
        previous_sunday,
        WEEKLY_END_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    return DateTimeRange(
        data_cut_type=DataCutType.WEEKLY,
        reference_date=reference_date,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
    )


def build_datetime_range(
    data_cut_type: str,
    reference_date: date,
) -> DateTimeRange:
    normalized_type = data_cut_type.strip().lower()

    try:
        selected_type = DataCutType(normalized_type)

    except ValueError as exc:
        raise ValueError(
            "data_cut_type은 'daily' 또는 "
            f"'weekly'여야 합니다. 입력값: {data_cut_type!r}"
        ) from exc

    if selected_type == DataCutType.DAILY:
        return build_daily_range(
            reference_date=reference_date,
        )

    return build_weekly_range(
        reference_date=reference_date,
    )


UNIX_EPOCH = datetime(
    1970,
    1,
    1,
    tzinfo=timezone.utc,
)


def datetime_to_epoch_ms(
    value: datetime,
) -> int:
    if value.tzinfo is None:
        raise ValueError(
            "timezone 정보가 없는 datetime입니다."
        )

    utc_value = value.astimezone(timezone.utc)
    delta = utc_value - UNIX_EPOCH

    return (
        delta.days * 86_400_000
        + delta.seconds * 1_000
        + delta.microseconds // 1_000
    )


def datetime_range_to_epoch_ms(
    date_range: DateTimeRange,
) -> tuple[int, int]:
    start_time_ms = datetime_to_epoch_ms(
        date_range.start_datetime
    )

    end_time_ms = datetime_to_epoch_ms(
        date_range.end_datetime
    )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "변환된 startTime이 endTime보다 작지 않습니다."
        )

    return start_time_ms, end_time_ms


def build_data_cut(
    data_cut_type: str,
    reference_date_text: str,
) -> tuple[DateTimeRange, int, int]:
    """
    main()에서 받은 data_cut_type과 기준 날짜를 이용해
    날짜 범위 및 epoch milliseconds를 생성한다.
    """
    reference_date = parse_reference_date(
        reference_date_text
    )

    date_range = build_datetime_range(
        data_cut_type=data_cut_type,
        reference_date=reference_date,
    )

    start_time_ms, end_time_ms = (
        datetime_range_to_epoch_ms(date_range)
    )

    print()
    print(
        "데이터 컷 유형:",
        date_range.data_cut_type.value,
    )
    print(
        "시작 일시:",
        date_range.start_datetime,
    )
    print(
        "종료 일시:",
        date_range.end_datetime,
    )
    print("startTime:", start_time_ms)
    print("endTime:", end_time_ms)

    return date_range, start_time_ms, end_time_ms


# =============================================================================
# 날짜 data cut이 반영된 payload template 생성
# =============================================================================

def build_date_payload_template(
    base_payload: dict[str, Any],
    start_time_ms: int,
    end_time_ms: int,
) -> dict[str, Any]:
    if not isinstance(base_payload, dict):
        raise TypeError(
            "base_payload는 dict여야 합니다."
        )

    if not isinstance(start_time_ms, int):
        raise TypeError(
            "start_time_ms는 int여야 합니다."
        )

    if not isinstance(end_time_ms, int):
        raise TypeError(
            "end_time_ms는 int여야 합니다."
        )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "startTime은 endTime보다 작아야 합니다."
        )

    run_payload_template = copy.deepcopy(
        base_payload
    )

    run_payload_template["startTime"] = start_time_ms
    run_payload_template["endTime"] = end_time_ms

    return run_payload_template


# =============================================================================
# Excel에서 # / Query 추출 및 Query payload 생성
# =============================================================================

@dataclass(frozen=True)
class CampaignQueryTask:
    row_id: int
    query: str

    @property
    def task_key(self) -> tuple[int, str]:
        return self.row_id, self.query


def normalize_row_id(
    value: Any,
) -> int:
    if value is None or pd.isna(value):
        raise ValueError(
            "# 값이 비어 있습니다."
        )

    if isinstance(value, bool):
        raise ValueError(
            "# 값은 boolean일 수 없습니다."
        )

    try:
        numeric_value = float(value)

    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"# 값을 숫자로 변환할 수 없습니다: {value!r}"
        ) from exc

    if not numeric_value.is_integer():
        raise ValueError(
            "# 값은 정수여야 합니다. "
            f"입력값: {value!r}"
        )

    row_id = int(numeric_value)

    if row_id <= 0:
        raise ValueError(
            "# 값은 1 이상의 정수여야 합니다. "
            f"입력값: {value!r}"
        )

    return row_id


EXCEL_ROW_ID_FORMULA_PATTERN = re.compile(
    r"^=\s*ROW\(\s*\)\s*-\s*ROW\(\s*"
    r"(?:(?:'[^']+'|[^!()]+)!)?"
    r"\$?[A-Z]{1,3}\$?(\d+)\s*\)\s*$",
    flags=re.IGNORECASE,
)


def resolve_excel_row_id(
    value: Any,
    excel_row_number: int,
) -> int:
    """
    Excel '#' 셀의 값을 실제 정수 ID로 변환한다.

    지원 형식:
    - 일반 숫자: 1, 2, 3 ...
    - 행 번호 수식: =ROW()-ROW($B$4)

    예를 들어 Excel 11행에서 수식이 =ROW()-ROW($B$4)이면
    row_id는 11 - 4 = 7로 계산한다.
    """
    if isinstance(value, str):
        formula = value.strip()

        if formula.startswith("="):
            formula_match = EXCEL_ROW_ID_FORMULA_PATTERN.fullmatch(
                formula
            )

            if formula_match is None:
                raise ValueError(
                    "지원하지 않는 # 수식입니다. "
                    "지원 형식 예시: =ROW()-ROW($B$4). "
                    f"입력값: {value!r}"
                )

            anchor_row_number = int(
                formula_match.group(1)
            )
            calculated_row_id = (
                excel_row_number - anchor_row_number
            )

            return normalize_row_id(
                calculated_row_id
            )

    return normalize_row_id(value)


def normalize_query(
    value: Any,
) -> str:
    if value is None or pd.isna(value):
        raise ValueError(
            "Query 값이 비어 있습니다."
        )

    query = str(value).strip()

    if not query:
        raise ValueError(
            "Query 값이 빈 문자열입니다."
        )

    # Query 내부 문법은 변경하지 않고 양끝 공백만 제거한다.
    return query


def normalize_header_name(
    value: Any,
) -> str:
    """
    Excel Header 비교용 문자열을 정규화한다.

    - 줄바꿈, 탭, 연속 공백을 일반 공백 1개로 통일
    - 앞뒤 공백 제거
    - 대소문자 차이 제거

    예시:
    "Buzz\nVolume" -> "buzz volume"
    "  Buzz   Volume  " -> "buzz volume"
    """
    if value is None:
        return ""

    normalized_value = (
        str(value)
        .replace("\u00A0", " ")
    )

    return " ".join(
        normalized_value.split()
    ).casefold()


def validate_campaign_query_tasks(
    tasks: list[CampaignQueryTask],
) -> None:
    if not tasks:
        raise ValueError(
            "처리할 캠페인 Query가 없습니다."
        )

    row_id_counts = Counter(
        task.row_id
        for task in tasks
    )

    duplicated_row_ids = [
        row_id
        for row_id, count in row_id_counts.items()
        if count > 1
    ]

    if duplicated_row_ids:
        raise ValueError(
            "# 컬럼에 중복 값이 있습니다: "
            f"{duplicated_row_ids[:20]}"
        )

    task_key_counts = Counter(
        task.task_key
        for task in tasks
    )

    duplicated_task_keys = [
        task_key
        for task_key, count in task_key_counts.items()
        if count > 1
    ]

    if duplicated_task_keys:
        raise ValueError(
            "(#, Query) 고유키가 중복되었습니다: "
            f"{duplicated_task_keys[:20]}"
        )


def load_campaign_query(
    excel_path: Path,
    sheet_name: str,
) -> list[CampaignQueryTask]:
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel 파일을 찾을 수 없습니다: {excel_path}"
        )

    if not excel_path.is_file():
        raise FileNotFoundError(
            f"Excel 경로가 파일이 아닙니다: {excel_path}"
        )

    campaign_df = pd.read_excel(
        excel_path,
        sheet_name=sheet_name,
        header=HEADER_ROW - 1,
        usecols="B:S",
    )

    required_columns = {
        ROW_ID_COLUMN_NAME,
        "Query",
    }

    missing_columns = (
        required_columns
        - set(campaign_df.columns)
    )

    if missing_columns:
        raise ValueError(
            "필수 컬럼이 없습니다: "
            f"{sorted(missing_columns)}"
        )

    tasks: list[CampaignQueryTask] = []
    validation_errors: list[str] = []
    skipped_empty_query_rows: list[int] = []

    for dataframe_index, row in campaign_df.iterrows():
        # HEADER_ROW 다음 행부터 데이터가 시작한다.
        excel_row_number = dataframe_index + HEADER_ROW + 1

        # Excel의 빈 셀은 pandas에서 NaN으로 읽힐 수 있다.
        # Query가 비어 있는 행은 API 호출 및 입력값 검증 대상에서 제외한다.
        raw_query = row["Query"]

        if (
            raw_query is None
            or pd.isna(raw_query)
            or str(raw_query).strip() == ""
        ):
            skipped_empty_query_rows.append(excel_row_number)
            continue

        try:
            row_id = normalize_row_id(
                row[ROW_ID_COLUMN_NAME]
            )

            query = normalize_query(
                raw_query
            )

            tasks.append(
                CampaignQueryTask(
                    row_id=row_id,
                    query=query,
                )
            )

        except ValueError as exc:
            validation_errors.append(
                f"Excel 행 {excel_row_number}: {exc}"
            )

    if validation_errors:
        error_preview = "\n".join(
            validation_errors[:20]
        )
        remaining_error_count = max(
            0,
            len(validation_errors) - 20,
        )
        extra_message = (
            f"\n외 {remaining_error_count}개 오류"
            if remaining_error_count
            else ""
        )

        raise ValueError(
            "Excel Query 입력값 검증에 실패했습니다.\n"
            f"{error_preview}{extra_message}"
        )

    validate_campaign_query_tasks(tasks)

    if skipped_empty_query_rows:
        print()
        print(
            "⏭️ Query가 비어 있어 건너뛴 Excel 행 수: "
            f"{len(skipped_empty_query_rows)}"
        )
        print(
            "건너뛴 Excel 행: "
            f"{skipped_empty_query_rows[:20]}"
        )

        if len(skipped_empty_query_rows) > 20:
            print(
                "추가로 건너뛴 행 수: "
                f"{len(skipped_empty_query_rows) - 20}"
            )

    return tasks

def build_query_payload(
    run_payload_template: dict[str, Any],
    query: Any,
) -> dict[str, Any]:
    query_text = normalize_query(query)

    payload = copy.deepcopy(
        run_payload_template
    )

    filters = payload.get("filters")

    if not isinstance(filters, list):
        raise ValueError(
            "payload의 filters가 리스트가 아닙니다."
        )

    query_filters = [
        filter_item
        for filter_item in filters
        if isinstance(filter_item, dict)
        and filter_item.get("dimensionName") == "QUERY"
    ]

    if len(query_filters) != 1:
        raise ValueError(
            "QUERY 필터가 정확히 1개여야 합니다. "
            f"현재 발견된 개수: {len(query_filters)}"
        )

    query_filter = query_filters[0]

    if query_filter.get("filterType") != "IN":
        raise ValueError(
            "QUERY 필터의 filterType이 'IN'이 아닙니다."
        )

    query_filter["values"] = [query_text]

    return payload

def fetch_sprinklr_data(
    base_url: str,
    endpoint: str,
    api_key: str | None,
    access_token: str | None,
    payload: dict[str, Any],
    session: requests.Session | None = None,
) -> dict[str, Any]:
    """Sprinklr Reporting API를 호출하고 JSON object 응답을 반환한다."""
    if not api_key:
        raise ValueError(
            "API_KEY가 없습니다. "
            "SPRINKLR_API_KEY 환경변수를 확인하세요."
        )

    if not access_token:
        raise ValueError(
            "ACCESS_TOKEN이 없습니다. "
            "SPRINKLR_ACCESS_TOKEN 환경변수를 확인하세요."
        )

    if not isinstance(payload, dict):
        raise TypeError(
            "payload는 dict여야 합니다."
        )

    url = (
        f"{base_url.rstrip('/')}/"
        f"{endpoint.lstrip('/')}"
    )

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Key": api_key,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    request_client = (
        session
        if session is not None
        else requests
    )

    try:
        response = request_client.post(
            url,
            headers=headers,
            json=payload,
            timeout=(10, 180),
        )

        response.raise_for_status()

    except requests.exceptions.Timeout as exc:
        raise RuntimeError(
            "Sprinklr API 요청 시간이 초과되었습니다.\n"
            f"URL: {url}"
        ) from exc

    except requests.exceptions.ConnectionError as exc:
        raise RuntimeError(
            "Sprinklr API 서버에 연결할 수 없습니다.\n"
            f"URL: {url}"
        ) from exc

    except requests.exceptions.HTTPError as exc:
        status_code = (
            exc.response.status_code
            if exc.response is not None
            else "unknown"
        )
        response_text = (
            exc.response.text[:3000]
            if exc.response is not None
            else ""
        )

        raise RuntimeError(
            "Sprinklr API가 HTTP 오류를 반환했습니다.\n"
            f"status_code: {status_code}\n"
            f"url: {url}\n"
            f"response: {response_text}"
        ) from exc

    except requests.exceptions.RequestException as exc:
        raise RuntimeError(
            "Sprinklr API 요청 중 오류가 발생했습니다.\n"
            f"URL: {url}"
        ) from exc

    try:
        response_json = response.json()

    except requests.exceptions.JSONDecodeError as exc:
        raise ValueError(
            "Sprinklr API 응답이 JSON 형식이 아닙니다.\n"
            f"status_code: {response.status_code}\n"
            f"response: {response.text[:3000]}"
        ) from exc

    if not isinstance(response_json, dict):
        raise TypeError(
            "Sprinklr API 응답의 최상위 구조는 "
            "JSON object여야 합니다. "
            f"실제 타입: {type(response_json).__name__}"
        )

    return response_json


def save_response_sample(
    response_json: dict[str, Any],
    widget_name: str,
    output_dir: str | Path = RESPONSE_SAMPLE_DIR,
) -> Path:
    safe_widget_name = (
        widget_name
        .replace(" ", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
    )

    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)

    file_path = path / f"{safe_widget_name}_response.json"

    with file_path.open(
        mode="w",
        encoding="utf-8",
    ) as file:
        json.dump(
            response_json,
            file,
            ensure_ascii=False,
            indent=2,
        )

    return file_path


def extract_mention_count(
    response_json: dict[str, Any],
) -> int:
    """
    다음 응답 구조에서 Mention 값을 추출한다.

    {
        "data": {
            "rows": [
                [265]
            ]
        }
    }
    """
    if not isinstance(response_json, dict):
        raise TypeError(
            "API 응답은 dict여야 합니다."
        )

    data = response_json.get("data")

    if not isinstance(data, dict):
        raise ValueError(
            "API 응답에 올바른 data object가 없습니다."
        )

    rows = data.get("rows")

    if not isinstance(rows, list):
        raise ValueError(
            "API 응답의 data.rows가 list가 아닙니다."
        )

    if not rows:
        raise ValueError(
            "API 응답의 data.rows가 비어 있습니다."
        )

    first_row = rows[0]

    if not isinstance(first_row, list):
        raise ValueError(
            "API 응답의 첫 번째 row가 list가 아닙니다."
        )

    if not first_row:
        raise ValueError(
            "API 응답의 첫 번째 row가 비어 있습니다."
        )

    raw_mention_count = first_row[0]

    if isinstance(raw_mention_count, bool):
        raise ValueError(
            "Mention 값이 boolean입니다."
        )

    try:
        numeric_mention_count = float(
            raw_mention_count
        )

    except (TypeError, ValueError) as exc:
        raise ValueError(
            "Mention 값을 숫자로 변환할 수 없습니다. "
            f"값: {raw_mention_count!r}"
        ) from exc

    if not numeric_mention_count.is_integer():
        raise ValueError(
            "Mention 값은 정수여야 합니다. "
            f"값: {raw_mention_count!r}"
        )

    mention_count = int(
        numeric_mention_count
    )

    if mention_count < 0:
        raise ValueError(
            "Mention 값은 0 이상이어야 합니다. "
            f"값: {mention_count}"
        )

    return mention_count


def update_mentions_in_excel(
    input_excel_path: Path,
    output_excel_path: Path,
    sheet_name: str,
    mentions_by_row_id: dict[int, int],
) -> None:
    """
    Excel의 '#' 값을 기준으로 Mention 결과를 매핑하여
    같은 행의 'Buzz Volume' 컬럼에 입력하고 새 파일로 저장한다.
    """
    if not input_excel_path.exists():
        raise FileNotFoundError(
            "입력 Excel 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if not input_excel_path.is_file():
        raise FileNotFoundError(
            "입력 Excel 경로가 파일이 아닙니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if input_excel_path.resolve() == output_excel_path.resolve():
        raise ValueError(
            "입력 Excel과 출력 Excel 경로가 같습니다. "
            "원본 보호를 위해 다른 경로를 사용해야 합니다."
        )

    if not mentions_by_row_id:
        raise ValueError(
            "Excel에 입력할 Mention 결과가 없습니다."
        )

    normalized_mentions: dict[int, int] = {}

    for raw_row_id, raw_mention_count in mentions_by_row_id.items():
        row_id = normalize_row_id(raw_row_id)

        if isinstance(raw_mention_count, bool):
            raise ValueError(
                "Mention 값은 boolean일 수 없습니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            )

        try:
            numeric_mention_count = float(raw_mention_count)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                "Mention 값을 숫자로 변환할 수 없습니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            ) from exc

        if not numeric_mention_count.is_integer():
            raise ValueError(
                "Mention 값은 정수여야 합니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            )

        mention_count = int(numeric_mention_count)

        if mention_count < 0:
            raise ValueError(
                "Mention 값은 0 이상이어야 합니다. "
                f"#={row_id}, 값={mention_count}"
            )

        normalized_mentions[row_id] = mention_count

    keep_vba = input_excel_path.suffix.lower() == ".xlsm"

    workbook = load_workbook(
        filename=input_excel_path,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "대상 시트를 찾을 수 없습니다.\n"
                f"대상 시트: {sheet_name}\n"
                f"현재 시트 목록: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]

        header_cells: dict[str, list[int]] = {
            ROW_ID_COLUMN_NAME: [],
            "Query": [],
            BUZZ_VOLUME_COLUMN_NAME: [],
        }

        # 실제 Excel Header는 "Buzz\nVolume"처럼 셀 내부 줄바꿈이
        # 포함될 수 있으므로, 비교할 때만 공백을 정규화한다.
        # Excel에 표시되는 원래 Header 값과 서식은 변경하지 않는다.
        normalized_header_lookup = {
            normalize_header_name(header_name): header_name
            for header_name in header_cells
        }

        for cell in worksheet[HEADER_ROW]:
            normalized_cell_header = normalize_header_name(
                cell.value
            )

            canonical_header_name = (
                normalized_header_lookup.get(
                    normalized_cell_header
                )
            )

            if canonical_header_name is not None:
                header_cells[canonical_header_name].append(
                    cell.column
                )

        if len(header_cells[ROW_ID_COLUMN_NAME]) != 1:
            raise ValueError(
                f"Header 행에서 {ROW_ID_COLUMN_NAME!r} 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells[ROW_ID_COLUMN_NAME])}"
            )

        if len(header_cells["Query"]) != 1:
            raise ValueError(
                "Header 행에서 'Query' 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells['Query'])}"
            )

        if len(header_cells[BUZZ_VOLUME_COLUMN_NAME]) != 1:
            raise ValueError(
                f"Header 행에서 {BUZZ_VOLUME_COLUMN_NAME!r} 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells[BUZZ_VOLUME_COLUMN_NAME])}"
            )

        row_id_column_number = (
            header_cells[ROW_ID_COLUMN_NAME][0]
        )
        query_column_number = (
            header_cells["Query"][0]
        )
        buzz_volume_column_number = (
            header_cells[BUZZ_VOLUME_COLUMN_NAME][0]
        )

        excel_row_by_id: dict[int, int] = {}
        invalid_row_id_errors: list[str] = []

        for excel_row_number in range(
            HEADER_ROW + 1,
            worksheet.max_row + 1,
        ):
            raw_query = worksheet.cell(
                row=excel_row_number,
                column=query_column_number,
            ).value

            # load_campaign_query()와 동일하게 Query가 빈 행은 무시한다.
            # 따라서 이 행의 # 값이 비어 있거나 다른 빈 Query 행과
            # 중복되어도 Buzz Volume 업데이트 대상에는 영향을 주지 않는다.
            if raw_query is None:
                continue

            if isinstance(raw_query, str) and not raw_query.strip():
                continue

            raw_row_id = worksheet.cell(
                row=excel_row_number,
                column=row_id_column_number,
            ).value

            try:
                row_id = resolve_excel_row_id(
                    value=raw_row_id,
                    excel_row_number=excel_row_number,
                )
            except ValueError as exc:
                invalid_row_id_errors.append(
                    f"Excel 행 {excel_row_number}: {exc}"
                )
                continue

            if row_id in excel_row_by_id:
                raise ValueError(
                    "Query가 입력된 Excel 행의 # 값이 "
                    "중복되었습니다.\n"
                    f"중복 # 값: {row_id}\n"
                    f"첫 번째 행: {excel_row_by_id[row_id]}\n"
                    f"두 번째 행: {excel_row_number}"
                )

            excel_row_by_id[row_id] = excel_row_number

        if invalid_row_id_errors:
            error_preview = "\n".join(invalid_row_id_errors[:20])
            remaining_error_count = max(
                0,
                len(invalid_row_id_errors) - 20,
            )
            extra_message = (
                f"\n외 {remaining_error_count}개 오류"
                if remaining_error_count
                else ""
            )

            raise ValueError(
                "Excel의 # 컬럼 검증에 실패했습니다.\n"
                f"{error_preview}{extra_message}"
            )

        missing_row_ids = sorted(
            set(normalized_mentions)
            - set(excel_row_by_id)
        )

        if missing_row_ids:
            raise ValueError(
                "Mention 결과의 # 값을 Excel에서 찾을 수 없습니다.\n"
                f"미매칭 # 값: {missing_row_ids}"
            )

        updated_count = 0
        newly_filled_count = 0
        overwritten_count = 0
        mention_total = 0

        for row_id, mention_count in normalized_mentions.items():
            excel_row_number = excel_row_by_id[row_id]

            buzz_volume_cell = worksheet.cell(
                row=excel_row_number,
                column=buzz_volume_column_number,
            )
            existing_buzz_volume = buzz_volume_cell.value

            if (
                existing_buzz_volume is None
                or (
                    isinstance(existing_buzz_volume, str)
                    and not existing_buzz_volume.strip()
                )
            ):
                newly_filled_count += 1
            else:
                overwritten_count += 1

            # 기존 셀이 비어 있으면 새 값을 입력하고,
            # 기존 숫자나 수식이 있으면 새 API 결과로 덮어쓴다.
            # 셀의 기존 서식은 openpyxl이 그대로 유지한다.
            buzz_volume_cell.value = mention_count

            updated_count += 1
            mention_total += mention_count

        if updated_count != len(normalized_mentions):
            raise RuntimeError(
                "Mention 업데이트 건수가 예상과 다릅니다.\n"
                f"예상 건수: {len(normalized_mentions)}\n"
                f"실제 건수: {updated_count}"
            )

        output_excel_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        temporary_output_path = output_excel_path.with_name(
            f".{output_excel_path.stem}"
            f"_temporary"
            f"{output_excel_path.suffix}"
        )

        try:
            workbook.save(temporary_output_path)
            temporary_output_path.replace(output_excel_path)
        finally:
            if temporary_output_path.exists():
                temporary_output_path.unlink()

        print()
        print("✅ Buzz Volume Excel 적재 완료")
        print(f"전체 업데이트 행 수: {updated_count}")
        print(f"빈 셀 신규 입력: {newly_filled_count}")
        print(f"기존 값 덮어쓰기: {overwritten_count}")
        print(f"Mention 합계: {mention_total:,}")
        print(f"결과 파일: {output_excel_path}")

    finally:
        workbook.close()


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    data_cut_type = input(
        "데이터 컷 유형을 입력하세요 "
        "(daily / weekly): "
    ).strip()

    reference_date_text = input(
        "기준 날짜를 입력하세요 "
        "(YYYY-MM-DD): "
    ).strip()

    date_range, start_time_ms, end_time_ms = build_data_cut(
        data_cut_type=data_cut_type,
        reference_date_text=reference_date_text,
    )

    input_excel_path, output_excel_path = build_excel_paths(
        reference_date_text
    )

    target_sheet_name = "로컬 캠페인 리스트_QHB8"
    base_payload_path = PAYLOAD_DIR / "buzz_volume_base_payload.json"

    base_payload = load_base_payload(base_payload_path)

    run_payload_template = build_date_payload_template(
        base_payload=base_payload,
        start_time_ms=start_time_ms,
        end_time_ms=end_time_ms,
    )

    campaign_tasks = load_campaign_query(
        excel_path=input_excel_path,
        sheet_name=target_sheet_name,
    )

    mentions_by_row_id: dict[int, int] = {}

    print()
    print(f"전체 API 호출 대상: {len(campaign_tasks)}개")

    with requests.Session() as session:
        for task_index, campaign_task in enumerate(
            campaign_tasks,
            start=1,
        ):
            print()
            print(
                f"[{task_index}/{len(campaign_tasks)}] "
                f"API 호출 시작: #={campaign_task.row_id}"
            )

            request_payload = build_query_payload(
                run_payload_template=run_payload_template,
                query=campaign_task.query,
            )

            response_json = fetch_sprinklr_data(
                base_url=SPRINKLR_BASE_URL,
                endpoint=ENDPOINT,
                api_key=API_KEY,
                access_token=ACCESS_TOKEN,
                payload=request_payload,
                session=session,
            )

            try:
                mention_count = extract_mention_count(
                    response_json=response_json,
                )

            except (TypeError, ValueError) as exc:
                if SAVE_FAILED_RESPONSES:
                    failed_response_path = save_response_sample(
                        response_json=response_json,
                        widget_name=(
                            f"failed_Buzz_Volume_row_"
                            f"{campaign_task.row_id}"
                        ),
                        output_dir=RESPONSE_SAMPLE_DIR,
                    )

                    print(
                        "⚠️ Mention 파싱 실패 응답 저장: "
                        f"{failed_response_path}"
                    )

                raise ValueError(
                    "Sprinklr 응답에서 Mention 값을 "
                    "추출하지 못했습니다.\n"
                    f"# 값: {campaign_task.row_id}\n"
                    f"Query: {campaign_task.query}"
                ) from exc

            if campaign_task.row_id in mentions_by_row_id:
                raise RuntimeError(
                    "Mention 결과 저장 중 중복된 # 값이 "
                    "발견되었습니다. "
                    f"#={campaign_task.row_id}"
                )

            mentions_by_row_id[campaign_task.row_id] = mention_count

            print(
                "API 호출 및 Mention 추출 완료: "
                f"#={campaign_task.row_id}, "
                f"Mentions={mention_count}"
            )

    expected_row_ids = {
        task.row_id
        for task in campaign_tasks
    }
    actual_row_ids = set(mentions_by_row_id)

    if actual_row_ids != expected_row_ids:
        missing_result_ids = sorted(
            expected_row_ids - actual_row_ids
        )
        unexpected_result_ids = sorted(
            actual_row_ids - expected_row_ids
        )

        raise RuntimeError(
            "API 결과와 Excel Query 대상의 # 값이 일치하지 않습니다.\n"
            f"누락 결과 # 값: {missing_result_ids}\n"
            f"예상하지 않은 # 값: {unexpected_result_ids}"
        )

    update_mentions_in_excel(
        input_excel_path=input_excel_path,
        output_excel_path=output_excel_path,
        sheet_name=target_sheet_name,
        mentions_by_row_id=mentions_by_row_id,
    )

    print()
    print("✅ 전체 작업 완료")
    print(
        f"데이터 컷: {date_range.start_datetime} "
        f"~ {date_range.end_datetime}"
    )
    print(
        f"전체 API 호출 완료: {len(mentions_by_row_id)}개"
    )
    print(
        f"결과 Excel: {output_excel_path}"
    )


if __name__ == "__main__":
    main()
