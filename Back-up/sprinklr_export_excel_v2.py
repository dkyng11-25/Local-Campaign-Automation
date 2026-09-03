# ============================================================
# 기존 Sprinklr to Excel Extraction에서 Profile URL column 추가
# ============================================================

import json
import os
from pathlib import Path

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import requests

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from dotenv import load_dotenv
import os

# =========================================================
# User Guide:
# Terminal에 python sprinklr_export_excel.py 기입 후
# 필요하신 scope에 해당하는 날짜/시간 기입 + Enter 하시면 실행 됩니다.
#
# 주의사항:
# 예를 들어서 7월 8일 18시부터 7월 9일 18시까지 조회하시는 경우
#
# 기입 형식: 
# start datetime: 2026-07-08 18:00:00
# end datetime: 2026-07-09 18:00:59
# 꼭 end datetime에는 시간에 59초까지 명시 해주셔야 합니다!!!
# =========================================================

# =========================================================
# 0. Widget configuration (Widget name & Payload Local Path)
# =========================================================

WIDGET_CONFIGS = [
    {
        "widget_name": "1.1. Comment 기준_Export용",
        "payload_path": r"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\payload\payload_1_1_comment.json",
    },

    {
        "widget_name": "1.2. Reply 기준_Export용",
        "payload_path": r"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\payload\payload_1_2_reply.json",
    },

    {
        "widget_name": "1.3. Repost 기준_Export용",
        "payload_path": r"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\payload\payload_1_3_repost.json",
    },

    {
        "widget_name": "2. 전략법인 전수조사 X",
        "payload_path": r"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\payload\payload_2_1_전략법인_X.json",
    },

    {
        "widget_name": "2. 전략법인 전수조사 IG",
        "payload_path": r"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\payload\payload_2_2_전략법인_IG.json"
    },
]

# =========================================================
# 1. 사용자 설정값 (URL, API Key, Access Token)
# =========================================================

"""사용자 설정 필요"""
SPRINKLR_BASE_URL = "https://api3.sprinklr.com/prod"
ENDPOINT = "/api/v2/reports/query"

# 보안상 실제 값은 코드에 직접 쓰기보다 환경변수로 관리
load_dotenv()
API_KEY = os.getenv("SPRINKLR_API_KEY")
ACCESS_TOKEN = os.getenv("SPRINKLR_ACCESS_TOKEN")

# =========================================================
# 2. 날짜/시간 문자열 변환 (milliseconds)
# =========================================================

def datetime_to_milliseconds(
        datetime_str: str,
        timezone_str: str = "Asia/Seoul"
        ) -> int:
    """
    사용자가 입력한 날짜/시간 문자열을 milliseconds로 변환

    입력예시: 
        "2024-06-01 00:00:00"

    출력예시:
        "1711920000000"
    """
    
    try:
        dt = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(
            f"Invalid datetime format: {datetime_str}."
            "Expected format is YYYY-MM-DD HH:MM:SS, e.g. 2026-07-01 00:00:00"
        )

    # 사용자가 입력한 시간을 지정 timezone 기준 시간으로 해석
    dt = dt.replace(tzinfo=ZoneInfo(timezone_str))

    # epoch seconds → milliseconds 변환
    epoch_ms = int(dt.timestamp() * 1000)

    return epoch_ms

def build_time_range_from_datetimes(
    start_datetime_str: str,
    end_datetime_str: str,
    timezone_str: str = "Asia/Seoul"
) -> tuple[int, int, str, str]:
    """
    사용자가 입력한 시작/종료 datetime 문자열을
    Sprinklr payload용 startTime/endTime milliseconds로 변환한다.

    입력 형식:
        YYYY-MM-DD HH:MM:SS

    입력 예시:
        start_datetime_str = "2026-07-07 18:30:00"
        end_datetime_str   = "2026-07-08 16:45:59"
    """

    try:
        datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S")
        datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(
            "Invalid datetime format. Expected format is YYYY-MM-DD HH:MM:SS, "
            "e.g. 2026-07-07 18:30:00"
        )

    start_time_ms = datetime_to_milliseconds(
        start_datetime_str,
        timezone_str=timezone_str
    )

    end_time_ms = datetime_to_milliseconds(
        end_datetime_str,
        timezone_str=timezone_str
    )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "Invalid time range. startTime must be earlier than endTime. "
            f"start={start_datetime_str}, end={end_datetime_str}"
        )

    return start_time_ms, end_time_ms, start_datetime_str, end_datetime_str

# =========================================================
# 3. Excel load or create
# =========================================================

def load_or_create_excel(excel_path: str) -> Workbook:
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)  # Ensure parent directories exist

    # Excel 파일이 이미 있으면 열기
    if path.exists():
        workbook = load_workbook(path)
    else:
        # 없으면 새 workbook 생성
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    
    return workbook

# =========================================================
# 4. Payload load & open
# =========================================================

def load_payload(payload_path: str) -> dict:
    path = Path(payload_path)

    if not path.exists():
        raise FileNotFoundError(f"Payload file not found: {payload_path}")
    
    if path.suffix.lower() != ".json":
        raise ValueError(f"Payload file must be a JSON file: {payload_path}")
    
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if not isinstance(payload, dict):
        raise ValueError(f"Payload JSON must be an dict but got {type(payload)}")

    return payload

# =========================================================
# 5. save modified payload (backup: optional)
# =========================================================

def save_payload(payload: dict, payload_path: str, make_backup: bool = False) -> None:
    """
    수정된 payload dict를 원래 JSON 파일에 저장

    make_backup=True이면 기존 payload 파일을 .bak 파일로 백업한 뒤 저장
    """

    path = Path(payload_path)

    if make_backup and path.exists():
        backup_path = path.with_suffix(path.suffix + ".bak")
        with open(path, "r", encoding="utf-8") as src:
            original_text = src.read()

        with open(backup_path, "w", encoding="utf-8") as bak:
            bak.write(original_text)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

# =========================================================
# 6. Payload 날짜 변경 (사용자 지정) 
# =========================================================

def update_payload_time_range(
        payload: dict,
        start_time_ms: int,
        end_time_ms: int
    ) -> dict:
    
    payload["startTime"] = start_time_ms
    payload["endTime"] = end_time_ms

    return payload

# =========================================================
# 7. Sprinklr API 호출
# =========================================================

def fetch_sprinklr_data(
    base_url: str,
    endpoint: str,
    api_key: str,
    access_token: str,
    payload: dict
) -> dict:
    if not api_key:
        raise ValueError("API_KEY is missing. Check SPRINKLR_API_KEY environment variable.")

    if not access_token:
        raise ValueError("ACCESS_TOKEN is missing. Check SPRINKLR_ACCESS_TOKEN environment variable.")

    url = f"{base_url}{endpoint}"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Key": api_key,
        "Content-Type": "application/json",
    }

    response = requests.post(
        url,
        headers=headers,
        json=payload,
        timeout=(10, 180)
    )

    # HTTP error 발생 시 여기서 에러 발생
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError:
        print("HTTP status:", response.status_code)
        print("Request URL:", url)
        print("Response text:")
        print(response.text[:3000])
        raise

    return response.json()

# =========================================================
# 8. Sprinklr response → DataFrame 변환
# =========================================================

def get_expected_columns_from_payload(payload: dict) -> list[str]:
    """
    payload의 groupBys/projections에서 예상 column 이름을 가져옴

    예:
        groupBys: POST_ID, ACCOUNT_ID
        projections: TOTAL_ENGAGEMENT

    결과:
        ["POST_ID", "ACCOUNT_ID", "TOTAL_ENGAGEMENT"]
    """
    columns = []

    for group_by in payload.get("groupBys", []):
        column_name = (
            group_by.get("heading")
            or group_by.get("dimensionName")
        )
        if column_name:
            columns.append(column_name)

    for projection in payload.get("projections", []):
        column_name = (
            projection.get("heading")
            or projection.get("measurementName")
        )
        if column_name:
            columns.append(column_name)

    return columns

# =========================================================
# 9. Sprinklr response에서 필요 row / heading 추출
# =========================================================
def find_rows_and_headings_in_response(response_json: dict):
    headings = None

    # Case 1. response_json["data"]가 dict인 경우
    # 예:
    # {
    #   "data": {
    #       "headings": [...],
    #       "rows": [...]
    #   }
    # }
    if isinstance(response_json, dict) and isinstance(response_json.get("data"), dict):
        data = response_json["data"]

        if isinstance(data.get("headings"), list):
            headings = data["headings"]

        if isinstance(data.get("rows"), list):
            return data["rows"], headings, "data.rows"

        if isinstance(data.get("results"), list):
            return data["results"], headings, "data.results"

        if isinstance(data.get("values"), list):
            return data["values"], headings, "data.values"

        if isinstance(data.get("data"), list):
            return data["data"], headings, "data.data"

        if headings is not None:
            return [], headings, "data.headings_only"

    # Case 2. response_json["data"] 자체가 list인 경우
    if isinstance(response_json, dict) and isinstance(response_json.get("data"), list):
        return response_json["data"], headings, "data"

    # Case 3. top-level rows
    if isinstance(response_json, dict) and isinstance(response_json.get("rows"), list):
        return response_json["rows"], headings, "rows"

    # Case 4. top-level results
    if isinstance(response_json, dict) and isinstance(response_json.get("results"), list):
        return response_json["results"], headings, "results"

    # Case 5. top-level headings only
    if isinstance(response_json, dict) and isinstance(response_json.get("headings"), list):
        headings = response_json["headings"]
        return [], headings, "headings_only"

    return None, headings, None

# =========================================================
# 10. 추출된 Created time + 9h
# =========================================================
def add_9_hours_to_created_time(value) -> str:
    """
    Sprinklr에서 추출된 Created Time 값에 9시간을 더한 뒤
    YYYY-MM-DD HH:MM:SS 형식으로 변환한다.

    처리 가능 입력:
        1. milliseconds timestamp
           예: 1783504943000

        2. Sprinklr 날짜 문자열
           예: "Jul 08, 2026, 08:00:25 PM"

    예:
        "Jul 08, 2026, 08:00:25 PM"
        -> "2026-07-09 05:00:25"
    """

    if value is None:
        return None

    # -----------------------------------------------------
    # Case 1. milliseconds timestamp인 경우
    # 예: 1783504943000
    # -----------------------------------------------------
    try:
        value_int = int(value)

        # milliseconds 기준으로 9시간 더하기
        adjusted_ms = value_int + (9 * 60 * 60 * 1000)

        dt = datetime.fromtimestamp(adjusted_ms / 1000, tz=ZoneInfo("UTC"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    except (ValueError, TypeError):
        pass

    # -----------------------------------------------------
    # Case 2. 문자열 날짜인 경우
    # 예: "Jul 08, 2026, 08:00:25 PM"
    # -----------------------------------------------------
    value_str = str(value).strip()

    possible_formats = [
        "%b %d, %Y, %I:%M:%S %p",  # Jul 08, 2026, 08:00:25 PM
        "%B %d, %Y, %I:%M:%S %p",  # July 08, 2026, 08:00:25 PM
    ]

    for fmt in possible_formats:
        try:
            dt = datetime.strptime(value_str, fmt)
            dt = dt + timedelta(hours=9)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    # 어떤 형식으로도 변환이 안 되면 원본 반환
    return value

# =========================================================
# 11. Conversation stream 관련 필드 추출
# =========================================================
def make_conversation_stream_dataframe(
    response_json: dict,
    target_sheet_name: str
) -> pd.DataFrame:
    """

    Raw Data_원문:
        - Conversation Stream
        - Permalink
        - Created Time
        - snType column

    Raw Data_전략법인:
        - Conversation Stream
        - Permalink
        - Created Time
        - snType column
        - Author Screen Name
    """

    rows, headings, found_path = find_rows_and_headings_in_response(response_json)

    if rows is None:
        raise ValueError(
            "Could not find rows in Sprinklr response. "
            "Please inspect saved response sample."
        )

    records = []

    for row in rows:
        message_obj = None

        if isinstance(row, list):
            for cell in row:
                if isinstance(cell, dict):
                    message_obj = cell
                    break

        elif isinstance(row, dict):
            message_obj = row

        if not isinstance(message_obj, dict):
            continue

        sender_profile = message_obj.get("senderProfile")

        if isinstance(sender_profile, dict):
            author_screen_name = sender_profile.get("name")
            profile_url = sender_profile.get("profileUrl")
            user_name = sender_profile.get("name")
        else:
            author_screen_name = None
            profile_url = None
            user_name = None

        record = {
            "Conversation Stream": message_obj.get("message"),
            "Campaign ID": message_obj.get("snMsgId"),
            "Profile URL": profile_url,
            "User Name": user_name,
            "Permalink": message_obj.get("permalink"),
            "Created Time": add_9_hours_to_created_time(message_obj.get("snCreatedTime")),
            "snType column": message_obj.get("snType"),
        }

        if target_sheet_name == "Raw Data_전략법인":
            record["Author Screen Name"] = author_screen_name

        records.append(record)

    if target_sheet_name == "Raw Data_전략법인":
        columns = [
            "Conversation Stream",
            "Campaign ID",
            "Profile URL",
            "User Name",
            "Permalink",
            "Created Time",
            "snType column",
            "Author Screen Name",
        ]
    else:
        columns = [
            "Conversation Stream",
            "Campaign ID",
            "Profile URL",
            "User Name",
            "Permalink",
            "Created Time",
            "snType column",
        ]

    df = pd.DataFrame(records, columns=columns)

    return df

# =========================================================
# 12. Sprinklr response에서 필요한 data 추출 
# =========================================================
def parse_sprinklr_response(response_json: dict, payload: dict | None = None) -> pd.DataFrame:
    print("Top-level response keys:", list(response_json.keys()))

    if "data" in response_json:
        print("response_json['data'] type:", type(response_json["data"]))
        if isinstance(response_json["data"], dict):
            print("data keys:", list(response_json["data"].keys()))

    if "errors" in response_json:
        print("Response errors:", response_json["errors"])

    rows, headings, found_path = find_rows_and_headings_in_response(response_json)

    if rows is None:
        raise ValueError(
            "Could not find row data or headings in Sprinklr response. "
            "Please inspect saved response sample."
        )

    print(f"Found response data at: {found_path}")
    print("headings:", headings)
    print("number of rows:", len(rows))

    # -----------------------------------------------------
    # Case 0. headings만 있고 rows가 없는 경우
    # -----------------------------------------------------
    if len(rows) == 0:
        if headings:
            print("No row data found. Returning empty DataFrame with headings.")
            return pd.DataFrame(columns=headings)

        print("No row data and no headings found. Returning empty DataFrame.")
        return pd.DataFrame()

    first_row = rows[0]
    print("first row type:", type(first_row))
    print("first row preview:", first_row)

    # -----------------------------------------------------
    # Case 1. rows = [{...}, {...}]
    # -----------------------------------------------------
    if isinstance(first_row, dict):
        df = pd.json_normalize(rows)

    # -----------------------------------------------------
    # Case 2 or 3. rows = [[...], [...]]
    # -----------------------------------------------------
    elif isinstance(first_row, list):

        # Case 3. rows = [[{...}], [{...}]]
        if len(first_row) == 1 and isinstance(first_row[0], dict):
            extracted_rows = []

            for row in rows:
                if isinstance(row, list) and len(row) == 1 and isinstance(row[0], dict):
                    extracted_rows.append(row[0])
                else:
                    extracted_rows.append({"value": row})

            df = pd.json_normalize(extracted_rows)

        # Case 2. rows = [["a", "b", 1], ["c", "d", 2]]
        else:
            df = pd.DataFrame(rows)

            if headings and len(headings) == len(df.columns):
                df.columns = headings
            elif payload is not None:
                expected_columns = get_expected_columns_from_payload(payload)

                if len(expected_columns) == len(df.columns):
                    df.columns = expected_columns
                else:
                    print(
                        "Warning: column count does not match. "
                        f"Headings: {len(headings) if headings else 0}, "
                        f"Payload columns: {len(expected_columns)}, "
                        f"Response columns: {len(df.columns)}."
                    )
            else:
                print("Warning: no headings or payload columns available.")

    else:
        df = pd.DataFrame({"value": rows})

    df = make_dataframe_excel_safe(df)

    return df

# =========================================================
# 13. Widget 이름 기반으로 excel sheet 이름 결정  
# =========================================================

def get_target_sheet_name(widget_name: str) -> str:
    """
    widget_name을 기반으로 Excel sheet 이름을 생성
    """
    widget_name = widget_name.strip()

    if widget_name.startswith("1."):
        return "Raw Data_원문"
    
    if widget_name.startswith("2."):
        return "Raw Data_전략법인"
   
    raise ValueError(
       f"Cannot determine target sheet for widget: {widget_name}."
       "Widget name must start with '1.' or '2.'."
   )

# =========================================================
# 14. Dict/list -> JSON string  
# =========================================================

def make_dataframe_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def convert_value(value):
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    return df.map(convert_value)

# =========================================================
# 15. DataFrame을 Excel 특정 sheet에 저장
# =========================================================

def append_dataframe_to_excel(
    workbook: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
) -> None:
    if not sheet_name:
        raise ValueError("Sheet_name cannot be empty.")

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    is_empty_sheet = (
        sheet.max_row == 1 
        and sheet.max_column == 1
        and sheet["A1"].value is None
    )

    write_header = is_empty_sheet 

    # DataFrame을 sheet에 쓰기
    for row in dataframe_to_rows(df, index=False, header=write_header):
        sheet.append(row)

# =========================================================
# 16. Srpinklr response 저장 파일 생성
# =========================================================
def save_response_sample(
    response_json: dict,
    widget_name: str,
    output_dir: str = "sprinklr_response_samples"
) -> None:
    safe_widget_name = (
        widget_name
        .replace(" ", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
    )

    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)

    file_path = path / f"{safe_widget_name}_response.json"

    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(response_json, f, ensure_ascii=False, indent=2)

# =========================================================
# 17. 실행 함수
# =========================================================

def main() -> None:

    # 1. User input 받기 및 가공
    print("=== Sprinklr Export to Excel ===")
    print("Please provide the following inputs:")

    start_datetime = input("Enter start datetime (YYYY-MM-DD HH:MM:SS): ").strip()
    end_datetime = input("Enter end datetime (YYYY-MM-DD HH:MM:SS): ").strip()

    naming_date = end_datetime.split(" ")[0].replace("-", "")
    naming_date = naming_date[2:8]
    
    OUTPUT_EXCEL_PATH = rf"C:\Users\KEARNEY\Desktop\Local_Campaign_Automation\output\{naming_date}_SLCC_SOV_Local Campaign Tracking_7월_v01.xlsx"
    
    start_time_ms, end_time_ms, data_cut_start, data_cut_end = build_time_range_from_datetimes(
        start_datetime_str=start_datetime,
        end_datetime_str=end_datetime,
        timezone_str="Asia/Seoul"
    )
 
    # 2. Date 저장할 excel 파일 생성 또는 load 
    workbook = load_or_create_excel(OUTPUT_EXCEL_PATH)

    # 3. 각 widget에 대한 payload load 및 update 
    for widget_config in WIDGET_CONFIGS:
        widget_name = widget_config["widget_name"]
        payload_path = widget_config["payload_path"]

        print(f"Processing widget: {widget_name}")

        print("Loading and Updating Sprinklr payload...")
        payload = load_payload(payload_path)
        
    
        payload = update_payload_time_range(
            payload=payload, 
            start_time_ms=start_time_ms, 
            end_time_ms=end_time_ms
        )

        save_payload(
            payload=payload,
            payload_path=payload_path,
            make_backup=False
        )
        
        # 4. 해당 widget에 대한 API 호출 
        print("Calling Sprinklr API...")
        response_json = fetch_sprinklr_data(
            base_url=SPRINKLR_BASE_URL,
            endpoint=ENDPOINT,
            api_key=API_KEY,
            access_token=ACCESS_TOKEN,
            payload=payload
        )

        # API response 저장
        save_response_sample(response_json, widget_name)

        # 5. API 요청에 대한 응답 dataframe으로 변환 
        TARGET_SHEET_NAME = get_target_sheet_name(widget_name)

        print("Converting response to DataFrame...")
        df = make_conversation_stream_dataframe(
            response_json=response_json,
            target_sheet_name=TARGET_SHEET_NAME
        )

        df = make_dataframe_excel_safe(df)

        df["source_widget"] = widget_name
        df["data_cut_start"] = payload["startTime"]
        df["data_cut_end"] = payload["endTime"]
        df["extracted_at"] = datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y-%m-%d %H:%M:%S")

        print("Writing data to Excel...")
        append_dataframe_to_excel(
            workbook=workbook,
            df=df,
            sheet_name=TARGET_SHEET_NAME
        )

    workbook.save(OUTPUT_EXCEL_PATH)

    print("Done.")
    print(f"Output file: {OUTPUT_EXCEL_PATH}")

if __name__ == "__main__":
    main()
